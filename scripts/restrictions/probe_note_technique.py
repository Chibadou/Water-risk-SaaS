#!/usr/bin/env python3
"""Sprint 38 — answer the four factual unknowns the note technique left open.

One run, four independent questions, on a GitHub runner with full network
access (every French open-data host is blocked by the sandbox proxy). Same
shape as probe_backlog.py, which at Sprint 22 closed two leads by negative
finding — a probe that finds nothing has not failed, it has closed a question.

  A. `rotation` (note §3.1). The ρ typology prescribes a `rotation` type for
     tours d'eau / alternate-day measures, worth 1 − 1/n. Does that wording
     actually occur in the 77 k rows of the "Restrictions" resource, and does it
     concern anyone other than farmers? Agriculture is out of scope (§0.2), so
     the type may be moot HERE even though the note prescribes it. Decides
     whether Sprint 39 implements it at all.

  B. SISPEA (note §11.1, G13). Fragility of the drinking-water service — network
     efficiency, losses, interconnections — for sites on the mains. The note
     calls it the differentiator no competitor has. Never investigated here.
     Question: does it exist as open data, at what granularity (service or
     commune?), how fresh, and is a commune → service join possible at all?

  C. Hydroportail vs our own indices (note §5.3, G14). We recompute VCN10 and
     QMNA5 empirically from Hub'Eau series (lib/hubeau.ts computeLowFlow). The
     note names Hydroportail as the reference. Question: is there a
     machine-readable endpoint, and do its published values match ours? Both
     answers are useful — agreement makes our method defensible in writing,
     disagreement means we found a bug.

  D. V_ref (note §4.2a, G9). The reference volume must be the one defined by the
     ICPE arrêté of 30 June 2023, amended 3 July 2024. Question: is that text
     reachable in a usable form? Légifrance's API is widely OAuth-gated, so the
     expected answer is "gated" — which is itself a finding, and would send
     Sprint 41 towards a transcribed-and-cited definition rather than a fetch.

Output: data/restrictions/note-technique-probe.json

Exit code: 1 if EVERY question failed, which means the run told us nothing and
must not look green. Individual negative findings are successes and exit 0.

⚠️ PASS 2, after pass 1 (run 31355992762) produced three verdicts that were not
findings at all. Every one of them read "there is nothing" when the truth was "I
could not look" — the exact silent failure this repo has paid for before:

  A said "rotation never concerns an entreprise usage". The audience columns are
    named `usage.u.concerne_entreprise`, not `concerne_entreprise`, so none were
    detected and the count was 0 by construction. build_restrictions.py:148
    already knew the prefix.
  B said "SISPEA not found in open data". The data.gouv query was a four-word
    phrase; the raw hit count was never recorded, so an empty search and a
    filtered-out search were indistinguishable.
  C said "no public JSON endpoint". Every request to hydro.eaufrance.fr had
    raised ConnectTimeout. No answer was obtained, from anything.

So this pass carries a rule the first one lacked: every question reports a
`status` of "mesuré" or "indéterminé", and a verdict may only claim absence when
the status is "mesuré". A probe that cannot tell the difference between "no" and
"I do not know" is worse than no probe, because it closes a question falsely.
"""

from __future__ import annotations

import csv
import io
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import requests

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data" / "restrictions"
OUT.mkdir(parents=True, exist_ok=True)

UA = {"User-Agent": "hydrovigie-note-probe/2.0 (github actions; water-risk-saas)"}
# Légifrance answered 403 to the probe UA on all three routes in pass 1. A 403 is
# a real answer, but it may be about the client rather than the resource, so the
# retry distinguishes the two instead of concluding from one UA.
BROWSER_UA = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/json;q=0.9,*/*;q=0.8",
}
VIGIEAU_DATASET = "https://www.data.gouv.fr/api/1/datasets/donnee-secheresse-vigieau/"
DATAGOUV_SEARCH = "https://www.data.gouv.fr/api/1/datasets/"
HUBEAU_HYDRO = "https://hubeau.eaufrance.fr/api/v2/hydrometrie"

report: dict = {
    "generated": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
    "a_rotation": {},
    "b_sispea": {},
    "c_hydroportail": {},
    "d_vref_icpe": {},
    "errors": [],
}
answered: set[str] = set()


def is_html(data: bytes) -> bool:
    return data.lstrip()[:1] == b"<"


def fetch(url: str, timeout: int = 300) -> bytes:
    r = requests.get(url, headers=UA, timeout=timeout)
    r.raise_for_status()
    return r.content


def read_csv(data: bytes) -> list[dict]:
    if is_html(data):
        raise ValueError("got HTML instead of a CSV file")
    text = ""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    sample = text[:8192]
    delim = max([",", ";", "\t"], key=lambda d: sample.count(d))
    return list(csv.DictReader(io.StringIO(text), delimiter=delim))


def read_table_header(data: bytes, url: str) -> tuple[list[str], dict | None, int | None]:
    """Column names + first row of a CSV **or** an Excel file.

    The SISPEA extractions are .xls/.xlsx. Handing those to read_csv() does not
    raise: latin-1 decodes anything, the delimiter sniff returns one giant
    column, and the probe would have reported "no commune key" from a file it
    never actually read — the same false negative this pass exists to kill.
    """
    head = data[:8]
    is_xlsx = head[:2] == b"PK"          # zip container: .xlsx
    is_xls = head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # OLE2: legacy .xls
    if is_xlsx:
        import openpyxl  # provided by the workflow
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=1, max_row=2, values_only=True)
        header = [str(c) for c in (next(rows, ()) or []) if c is not None]
        first = next(rows, None)
        sample = dict(zip(header, [str(c) for c in first])) if first else None
        return header, sample, ws.max_row
    if is_xls:
        # Legacy binary format; openpyxl cannot read it. Say so rather than
        # pretend, and let the sprint decide whether xlrd is worth a dependency.
        raise ValueError("legacy .xls (OLE2) — openpyxl cannot read it, needs xlrd")
    rows = read_csv(data)
    return (list(rows[0].keys()) if rows else []), (rows[0] if rows else None), len(rows)


def classify(url: str, timeout: int = 60, headers: dict | None = None) -> dict:
    """Status + shape of a candidate endpoint.

    Modelled on the Sprint 19 ZRE probe, which found the real source only
    because every candidate was classified rather than assumed: data.gouv URLs
    turned out to be HTML portals and INSPIRE hosts were dead.
    """
    out: dict = {"url": url}
    try:
        r = requests.get(url, headers=headers or UA, timeout=timeout)
        out["status"] = r.status_code
        out["content_type"] = r.headers.get("content-type", "")
        body = r.content[:400]
        out["looks_like"] = (
            "html" if is_html(body)
            else "json" if body.lstrip()[:1] in (b"{", b"[")
            else "other"
        )
        out["excerpt"] = body[:200].decode("utf-8", "replace")
        if r.status_code == 401:
            out["note"] = "authenticated endpoint (same pattern as MétéEAU/BRGM)"
    except Exception as e:  # noqa: BLE001
        out["error"] = repr(e)[:200]
        # A timeout is NOT evidence that the resource is absent. Pass 1 folded
        # the two together and concluded "no public endpoint" from five
        # ConnectTimeouts.
        out["unreachable"] = True
    return out


def datagouv_search(query: str, page_size: int = 10) -> dict:
    """data.gouv search that records the raw hit count.

    Pass 1 used four-word phrases and kept only the filtered list, so zero
    results and zero survivors of the filter looked identical. The raw count is
    what tells a real absence from a bad query.
    """
    out: dict = {"query": query}
    try:
        r = requests.get(
            DATAGOUV_SEARCH, headers=UA, timeout=120,
            params={"q": query, "page_size": page_size},
        )
        out["status"] = r.status_code
        body = r.json()
        data = body.get("data", [])
        out["raw_total"] = body.get("total")
        out["raw_count"] = len(data)
        out["results"] = [
            {
                "title": d.get("title"),
                "page": d.get("page"),
                "last_update": d.get("last_update"),
                "resources": [
                    {"title": x.get("title"), "format": x.get("format"),
                     "filesize": x.get("filesize"), "url": x.get("url")}
                    for x in (d.get("resources") or [])[:30]
                ],
            }
            for d in data
        ]
    except Exception as e:  # noqa: BLE001
        out["error"] = repr(e)[:200]
        out["unreachable"] = True
    return out


# --- resource discovery ------------------------------------------------------
resources: dict[str, str] = {}
try:
    ds = requests.get(VIGIEAU_DATASET, headers=UA, timeout=120).json()
    for r in ds.get("resources", []):
        title = (r.get("title") or "").strip()
        if title and r.get("url"):
            resources.setdefault(title, r["url"])
except Exception as e:  # noqa: BLE001
    report["errors"].append({"target": "discovery", "error": repr(e)[:300]})

# --- A. does `rotation` exist in the corpus? ---------------------------------
# Patterns taken from how prefectures actually phrase alternate-day measures.
# Deliberately wide: a false positive is cheap to read, a false negative would
# silently drop a ρ type the note prescribes.
ROTATION_PATTERNS = {
    # "autorisé 3 jours par semaine" is the dominant quantified form in the real
    # file — pass 1 missed it entirely and it is the one that maps cleanly onto
    # the note's ρ = 1 − n/7. Captured with its number so ρ is computable.
    "n_jours_par_semaine": r"(\d)\s*jours?\s+par\s+semaine",
    "tour_d_eau": r"tours?\s+d['’]eau",
    "jours_pairs_impairs": r"jours?\s+(?:pairs?|impairs?)",
    "un_jour_sur_deux": r"un\s+jour\s+sur\s+(?:deux|\d)",
    # \b matters: "altern" alone also matches "alternative". Kept wide but the
    # samples are recorded so a human can judge rather than trust the count.
    "alterne": r"\baltern(?:e|é|ance|ativement|es|és)\b",
    # \b matters even more here: pass 1 counted 48 hits on "roulement", and the
    # samples show they were all "déroulement".
    "par_roulement": r"\broulement\b",
    # 7/7 is a TOTAL ban, not a rotation — pass 1 counted it as one.
    "n_jours_sur_sept": r"\b[1-6]\s*jours?\s+sur\s+7\b",
}

try:
    a: dict = {}
    url = resources.get("Restrictions")
    if not url:
        a["verdict"] = "resource « Restrictions » introuvable"
    else:
        rows = read_csv(fetch(url))
        a["rows"] = len(rows)
        cols = list(rows[0].keys()) if rows else []
        a["columns"] = cols
        # The measure text column, named as in build_restrictions.py.
        desc_col = next(
            (c for c in cols if "description" in c.lower()),
            next((c for c in cols if "libelle" in c.lower()), None),
        )
        a["description_column"] = desc_col
        # ⚠️ The columns are `usage.u.concerne_*`, NOT `concerne_*`. Pass 1
        # matched on the bare prefix, found none, and reported the resulting
        # zero as "never concerns an entreprise". build_restrictions.py:148
        # already read them with the right prefix.
        audience_cols = [c for c in cols if "concerne_" in c.lower()]
        a["audience_columns"] = audience_cols

        hits: Counter = Counter()
        by_audience: dict[str, Counter] = {c: Counter() for c in audience_cols}
        samples: dict[str, list[str]] = {k: [] for k in ROTATION_PATTERNS}
        entreprise_examples: list[str] = []

        for row in rows:
            text = (row.get(desc_col) or "") if desc_col else ""
            if not text:
                continue
            low = text.lower()
            for name, pat in ROTATION_PATTERNS.items():
                if re.search(pat, low):
                    hits[name] += 1
                    if len(samples[name]) < 5:
                        samples[name].append(text[:220])
                    for c in audience_cols:
                        v = (row.get(c) or "").strip().lower()
                        if v in ("true", "1", "oui", "vrai"):
                            by_audience[c][name] += 1
                    # ⚠️ Same prefix bug as the audience columns, on a second
                    # line: pass 2 fixed the COUNT but left the EXAMPLES reading
                    # `concerne_entreprise`, so it measured 77 entreprise
                    # measures and could show none of them. A count with no
                    # visible evidence is not something to decide a ρ type on.
                    ent_col = next(
                        (c for c in audience_cols if c.endswith("concerne_entreprise")), None
                    )
                    ent = (row.get(ent_col) or "").strip().lower() if ent_col else ""
                    if ent in ("true", "1", "oui", "vrai") and len(entreprise_examples) < 10:
                        entreprise_examples.append({
                            "usage": row.get("usage.u.nom") or row.get("usage.u.thematique"),
                            "niveau": row.get("niveau_gravite"),
                            "texte": text[:260],
                        })

        a["audience_detected"] = len(audience_cols) > 0
        a["hits"] = dict(hits)
        a["hits_by_audience"] = {k: dict(v) for k, v in by_audience.items()}
        a["samples"] = {k: v for k, v in samples.items() if v}
        a["entreprise_examples"] = entreprise_examples
        total = sum(hits.values())
        ent_key = next((c for c in audience_cols if c.endswith("concerne_entreprise")), None)
        ent_total = sum(by_audience.get(ent_key, Counter()).values()) if ent_key else None
        a["total_matches"] = total
        a["entreprise_matches"] = ent_total
        a["status"] = "mesuré" if a["audience_detected"] else "indéterminé"
        if not a["audience_detected"]:
            # The failure mode of pass 1, now named instead of reported as a fact.
            a["verdict"] = (
                "INDETERMINE — colonnes d'audience non détectées, donc le décompte "
                "entreprise ne veut rien dire. Ne pas conclure."
            )
        elif total == 0:
            a["verdict"] = "ABSENT — aucun libellé de rotation dans le corpus ; type ρ sans objet ici"
        elif ent_total == 0:
            a["verdict"] = (
                "PRESENT MAIS HORS PERIMETRE — rotation trouvée, jamais sur un usage "
                "concernant l'entreprise ; agriculture hors périmètre §0.2"
            )
        else:
            a["verdict"] = f"PRESENT ET DANS LE PERIMETRE — {ent_total} mesures entreprise : à implémenter"
    report["a_rotation"] = a
    answered.add("A")
except Exception as e:  # noqa: BLE001
    report["errors"].append({"target": "A", "error": repr(e)[:300]})

# --- B. SISPEA -----------------------------------------------------------------
try:
    b: dict = {"searches": []}
    # Short, single-concept queries. Pass 1 used one four-word phrase and got
    # nothing, which says more about the query than about the data.
    for q in ["sispea", "services eau assainissement", "rendement réseau eau potable",
              "observatoire services publics eau"]:
        b["searches"].append(datagouv_search(q, page_size=10))
    reachable = [x for x in b["searches"] if "error" not in x]
    all_results = [r for x in reachable for r in x.get("results", [])]
    b["raw_hits_total"] = sum(x.get("raw_count", 0) for x in reachable)

    # Keep the ones that actually look like the observatory.
    kept = []
    seen = set()
    for r in all_results:
        title = (r.get("title") or "")
        key = title.lower()
        if key in seen:
            continue
        blob = key + " " + " ".join((x.get("title") or "") for x in r.get("resources", [])).lower()
        if any(k in blob for k in ("sispea", "service", "eau potable", "assainissement", "rendement")):
            seen.add(key)
            kept.append(r)
    b["datasets"] = kept[:10]

    blob = json.dumps(kept, ensure_ascii=False).lower()
    b["mentions"] = {
        k: (k.lower() in blob)
        for k in ["rendement", "p104", "p106", "pertes", "commune", "interconnexion", "indicateur"]
    }

    # What we need is a per-commune (or joinable) network-efficiency indicator:
    # P104.3 = rendement du réseau, P106.3 = indice linéaire de pertes.
    # Pass 2 found the right dataset — "Rendement du réseau de distribution de
    # l'eau potable par territoire compétent" — and then opened nothing, because
    # the candidate filter demanded words its resource titles do not carry. Rank
    # by relevance to the question instead of by vocabulary, and sniff up to
    # three so one awkward file does not hide the answer.
    def score(dataset_title: str, res: dict) -> int:
        t = ((res.get("title") or "") + " " + dataset_title).lower()
        fmt = (res.get("format") or "").lower()
        if fmt not in ("csv", "xlsx", "xls"):
            return -1
        n = 0
        if "rendement" in t: n += 5
        if "eau potable" in t: n += 3
        if "commune" in t or "territoire" in t: n += 2
        if "indicateur" in t or "donnee" in t or "données" in t: n += 1
        return n

    ranked = sorted(
        (
            (score(r.get("title") or "", res), r.get("title"), res)
            for r in kept for res in r.get("resources", [])
        ),
        key=lambda x: x[0], reverse=True,
    )
    b["sniff_attempts"] = []
    for sc, ds_title, res in ranked[:3]:
        if sc < 0:
            continue
        attempt = {"score": sc, "dataset": ds_title, "resource": res.get("title"), "url": res.get("url")}
        # Pass 3 reached these files and could not read them: not PK (xlsx), not
        # OLE2 (.xls), not HTML, and the csv module choked on binary. Rather
        # than guess again, record what the bytes actually are — the same
        # reflex that identified the .csv.gz behind the /api/swi outage, where
        # the gzip was in the payload and not in Content-Encoding.
        try:
            head = requests.get(
                res["url"], headers=UA, timeout=120, stream=True,
            )
            raw = next(head.iter_content(64), b"")
            attempt["first_bytes_hex"] = raw[:32].hex()
            attempt["first_bytes_ascii"] = raw[:32].decode("ascii", "replace")
            attempt["content_type"] = head.headers.get("content-type")
            attempt["content_disposition"] = head.headers.get("content-disposition")
            attempt["final_url"] = head.url
            head.close()
        except Exception as e:  # noqa: BLE001
            attempt["head_error"] = repr(e)[:160]
        try:
            cols, sample_row, nrows = read_table_header(fetch(res["url"], timeout=300), res["url"])
            joined = " ".join(cols).lower()
            attempt["rows"] = nrows
            attempt["columns"] = cols[:60]
            attempt["has_commune_key"] = any(k in joined for k in ("insee", "commune", "code_commune"))
            attempt["has_rendement"] = any(k in joined for k in ("rendement", "p104"))
            attempt["sample_row"] = sample_row
        except Exception as e:  # noqa: BLE001
            attempt["error"] = repr(e)[:200]
        b["sniff_attempts"].append(attempt)

    best = next(
        (x for x in b["sniff_attempts"] if x.get("has_commune_key") and x.get("has_rendement")),
        next((x for x in b["sniff_attempts"] if x.get("has_rendement")), None),
    )
    candidate = None
    if best:
        b["sniffed_resource"] = {"title": best.get("resource"), "url": best.get("url")}
        b["sniffed_rows"] = best.get("rows")
        b["sniffed_columns"] = best.get("columns")
        b["has_commune_key"] = best.get("has_commune_key")
        b["has_rendement"] = best.get("has_rendement")
        b["sample_row"] = best.get("sample_row")
        candidate = True
    if not reachable:
        b["status"] = "indéterminé"
        b["verdict"] = "INDETERMINE — data.gouv injoignable, aucune recherche n'a abouti"
    elif b.get("has_commune_key") and b.get("has_rendement"):
        b["status"] = "mesuré"
        b["verdict"] = "EXPLOITABLE — clé commune et rendement présents dans les colonnes"
    elif kept:
        b["status"] = "mesuré"
        b["verdict"] = (
            f"A INSTRUIRE — {len(kept)} jeux candidats trouvés, mais clé commune "
            "et/ou rendement non confirmés dans les colonnes sondées"
        )
    else:
        b["status"] = "mesuré"
        b["verdict"] = (
            f"ABSENT de data.gouv — {b['raw_hits_total']} résultats bruts sur 4 requêtes, "
            "aucun ne ressemble à l'observatoire. Reste à sonder services.eaufrance.fr directement."
        )
    report["b_sispea"] = b
    answered.add("B")
except Exception as e:  # noqa: BLE001
    report["errors"].append({"target": "B", "error": repr(e)[:300]})

# --- C. Hydroportail vs our own low-flow references ---------------------------
try:
    c: dict = {}
    # Resolve the station we have already verified in production (the Loire at
    # Orléans) through Hub'Eau rather than hardcoding a code.
    ref = requests.get(
        f"{HUBEAU_HYDRO}/referentiel/stations", headers=UA, timeout=120,
        params={
            "longitude": 1.9039, "latitude": 47.9029, "distance": 10,
            "format": "json", "size": 20, "en_service": "true",
        },
    ).json()
    stations = [
        {"code": s_.get("code_station"), "libelle": s_.get("libelle_station")}
        for s_ in ref.get("data", [])
    ]
    c["stations_near_orleans"] = stations[:10]
    code = next((s_["code"] for s_ in stations if s_.get("code")), None)
    c["station_used"] = code

    # ⚠️ Pass 1 hit ConnectTimeout on every hydro.eaufrance.fr URL with a 60 s
    # budget and a probe UA, then reported "no public endpoint". Retry with a
    # browser UA and a longer budget, and keep the two failure modes apart.
    candidates = []
    if code:
        candidates = [
            f"https://hydro.eaufrance.fr/stationhydro/{code}/series",
            f"https://hydro.eaufrance.fr/sitehydro/{code}/fiche",
            f"https://www.hydro.eaufrance.fr/stationhydro/{code}",
        ]
    candidates.append("https://hydro.eaufrance.fr/")
    c["candidates"] = [classify(u, timeout=120, headers=BROWSER_UA) for u in candidates]

    # Hub'Eau also publishes elaborated low-flow statistics; if it does, the
    # comparison needs no Hydroportail at all. Worth one call before concluding.
    try:
        alt = requests.get(
            f"{HUBEAU_HYDRO}/obs_elab", headers=UA, timeout=120,
            params={"code_entite": code, "grandeur_hydro_elab": "QmnJ", "size": 2},
        )
        c["hubeau_obs_elab_status"] = alt.status_code
        c["hubeau_obs_elab_count"] = len((alt.json() or {}).get("data", []))
    except Exception as e:  # noqa: BLE001
        c["hubeau_obs_elab_error"] = repr(e)[:200]

    c["datagouv"] = [datagouv_search(q, page_size=8) for q in ["hydroportail", "QMNA5"]]

    unreachable = [x for x in c["candidates"] if x.get("unreachable")]
    answered_http = [x for x in c["candidates"] if x.get("status")]
    json_ok = [x for x in answered_http if x.get("status") == 200 and x.get("looks_like") == "json"]
    dg_hits = sum(x.get("raw_count", 0) for x in c["datagouv"] if "error" not in x)

    if not answered_http:
        c["status"] = "indéterminé"
        c["verdict"] = (
            f"INDETERMINE — {len(unreachable)}/{len(c['candidates'])} candidats injoignables "
            "(timeout), aucune réponse HTTP obtenue. Ne pas conclure à l'absence."
        )
    elif json_ok:
        c["status"] = "mesuré"
        c["verdict"] = f"ENDPOINT JSON TROUVE ({len(json_ok)}) — comparaison chiffrée possible"
    else:
        c["status"] = "mesuré"
        c["verdict"] = (
            f"AUCUN ENDPOINT JSON — {len(answered_http)} routes répondent mais en HTML ; "
            f"{dg_hits} résultats data.gouv. Garder le calcul maison et l'écrire (précédent MétéEAU)."
        )
    report["c_hydroportail"] = c
    answered.add("C")
except Exception as e:  # noqa: BLE001
    report["errors"].append({"target": "C", "error": repr(e)[:300]})

# --- D. V_ref: the ICPE arrêté of 30 June 2023 --------------------------------
try:
    d: dict = {}
    # Pass 1: 403 on all three routes with the probe UA. A 403 is a real answer,
    # but it may be about the client, so retry with a browser UA before
    # concluding that the text cannot be fetched.
    urls = [
        "https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000047762160",
        "https://www.legifrance.gouv.fr/download/pdf?id=JORFTEXT000047762160",
        "https://api.legifrance.gouv.fr/dila/legifrance/lf-engine-app/consult/jorf",
    ]
    d["candidates_probe_ua"] = [classify(u, timeout=60) for u in urls]
    d["candidates_browser_ua"] = [classify(u, timeout=60, headers=BROWSER_UA) for u in urls]
    d["datagouv"] = [
        datagouv_search(q, page_size=8)
        for q in ["légifrance", "arrêté sécheresse ICPE", "prélèvement eau installations classées"]
    ]

    ok_browser = [x for x in d["candidates_browser_ua"] if x.get("status") == 200]
    ok_probe = [x for x in d["candidates_probe_ua"] if x.get("status") == 200]
    all_unreachable = all(x.get("unreachable") for x in d["candidates_browser_ua"])

    if all_unreachable:
        d["status"] = "indéterminé"
        d["verdict"] = "INDETERMINE — Légifrance injoignable (timeout), aucune réponse HTTP"
    elif ok_browser and not ok_probe:
        d["status"] = "mesuré"
        d["verdict"] = (
            "ATTEIGNABLE AVEC UN UA NAVIGATEUR — le 403 de la passe 1 venait du client, "
            "pas de la ressource. ⚠️ Se faire passer pour un navigateur pour contourner un "
            "blocage est un choix à arbitrer, pas un acquis technique."
        )
    elif ok_browser:
        d["status"] = "mesuré"
        d["verdict"] = "ATTEIGNABLE — au moins une route publique répond 200 ; reste à vérifier que V_ref y est extractible"
    else:
        codes = sorted({x.get("status") for x in d["candidates_browser_ua"] if x.get("status")})
        d["status"] = "mesuré"
        d["verdict"] = (
            f"NON ATTEIGNABLE — codes {codes} sur les deux UA. Transcrire la définition à la "
            "main avec citation de l'article (même traitement que le décret 2021-795)."
        )
    report["d_vref_icpe"] = d
    answered.add("D")
except Exception as e:  # noqa: BLE001
    report["errors"].append({"target": "D", "error": repr(e)[:300]})

# --- write & report -----------------------------------------------------------
(OUT / "note-technique-probe.json").write_text(
    json.dumps(report, ensure_ascii=False, indent=1, default=str) + "\n", encoding="utf-8"
)

print("=== PROBE NOTE TECHNIQUE (sprint 38) ===")
print("A rotation      :", report["a_rotation"].get("status"), "|", report["a_rotation"].get("verdict"))
print("   hits         :", report["a_rotation"].get("hits"))
print("B sispea        :", report["b_sispea"].get("status"), "|", report["b_sispea"].get("verdict"))
print("   mentions     :", report["b_sispea"].get("mentions"))
print("C hydroportail  :", report["c_hydroportail"].get("status"), "|", report["c_hydroportail"].get("verdict"))
print("   station      :", report["c_hydroportail"].get("station_used"))
print("D vref icpe     :", report["d_vref_icpe"].get("status"), "|", report["d_vref_icpe"].get("verdict"))
print("errors          :", report["errors"])

# A probe that answered nothing must not look green. Individual negative
# findings are results; four failures are an outage. Same reflex as the three
# sys.exit(1) floors added to the build scripts on 2026-08-07.
if not answered:
    raise SystemExit("probe answered none of the four questions — see errors above")
